import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO

st.set_page_config(page_title="SKU Analysis & Shelf Planner", layout="wide")

st.title("📊 SKU Performance + Shelf Capacity Planner")

st.write("Upload your SKU file, get performance-based recommendations (Expand, Retain, Delist), and check if your shelf space can accommodate the suggested facings.")

# --- File uploader ---
uploaded_file = st.file_uploader("📂 Upload CSV file", type=["csv"])

if uploaded_file is not None:
    df = pd.read_csv(uploaded_file)

    # --- Normalization & Scoring ---
    df['Sales_Norm'] = df['Sales'] / df['Sales'].max()
    df['Volume_Norm'] = df['Volume'] / df['Volume'].max()
    df['Margin_Norm'] = df['Margin'] / df['Margin'].max()

    df['Score'] = (df['Sales_Norm'] * 0.30) + (df['Volume_Norm'] * 0.30) + (df['Margin_Norm'] * 0.40)

    cutoff_expand = df['Score'].quantile(0.70)
    cutoff_delist = df['Score'].quantile(0.30)

    def classify(score):
        if score >= cutoff_expand:
            return "Expand"
        elif score <= cutoff_delist:
            return "Delist"
        else:
            return "Retain"

    df['Recommendation'] = df['Score'].apply(classify)

    def explain(row):
        if row['Recommendation'] == "Expand":
            return "High sales, volume, or margin → Increase facings or distribution."
        elif row['Recommendation'] == "Delist":
            return "Low performance → Candidate for phase-out."
        else:
            return "Balanced performance → Maintain current space."

    def move_out_plan(row):
        return "Consider promo bundling, discounting, or supplier return." if row['Recommendation'] == "Delist" else "-"

    df['Explanation'] = df.apply(explain, axis=1)
    df['Move-Out Plan'] = df.apply(move_out_plan, axis=1)

    # --- Shelf space planner ---
    st.sidebar.header("📏 Shelf Space Inputs")
    shelf_space = st.sidebar.number_input("Total Shelf Space (inches)", min_value=0.0, value=0.0, step=0.5)

    if 'Width_in' not in df.columns:
        st.sidebar.warning("No 'Width_in' column detected in CSV. Add SKU width (inches) to enable capacity calculation.")
        df['Width_in'] = 0

    # Suggested facings based on recommendation
    facing_map = {"Expand": 3, "Retain": 2, "Delist": 1}
    df['Suggested Facings'] = df['Recommendation'].map(facing_map)

    df['Space_Required'] = df['Width_in'] * df['Suggested Facings']

    total_space_required = df['Space_Required'].sum()

    if shelf_space > 0:
        st.sidebar.metric(label="Total Required Space (in)", value=round(total_space_required, 2))
        if total_space_required > shelf_space:
            st.sidebar.error("❌ Not enough shelf space! Consider reducing Delist SKUs first.")
        else:
            st.sidebar.success("✅ Fits within shelf capacity.")

    # --- Display detailed results ---
    st.subheader("📋 Detailed Results")

    def color_table(val):
        if val == "Expand":
            return "background-color: #c6efce"  # light green
        elif val == "Delist":
            return "background-color: #ffc7ce"  # light red
        elif val == "Retain":
            return "background-color: #ffeb9c"  # light yellow
        return ""

    st.dataframe(df.style.applymap(color_table, subset=["Recommendation"]), use_container_width=True)

    # --- Summary Chart ---
    st.subheader("📊 Summary of Recommendations")
    summary = df['Recommendation'].value_counts()
    fig, ax = plt.subplots()
    summary.plot(kind='bar', color=["#c6efce", "#ffeb9c", "#ffc7ce"], ax=ax)
    ax.set_title("Recommendation Breakdown")
    ax.set_ylabel("Number of SKUs")
    st.pyplot(fig)

    # --- Downloadable Results ---
    df_export = df.copy()

    def to_excel_bytes(df_export):
        for engine in ('xlsxwriter', 'openpyxl'):
            output = BytesIO()
            try:
                with pd.ExcelWriter(output, engine=engine) as writer:
                    df_export.to_excel(writer, index=False, sheet_name='SKU_Recommendations')
                    workbook = writer.book
                    worksheet = writer.sheets['SKU_Recommendations']

                    if engine == 'xlsxwriter':
                        fmt_expand = workbook.add_format({'bg_color': '#c6efce'})
                        fmt_retain = workbook.add_format({'bg_color': '#ffeb9c'})
                        fmt_delist = workbook.add_format({'bg_color': '#ffc7ce'})

                        headers = df_export.columns.tolist()
                        if 'Recommendation' in headers:
                            rec_col_idx = headers.index('Recommendation')
                            for row_num, rec_val in enumerate(df_export['Recommendation'], start=1):
                                if rec_val == 'Expand':
                                    worksheet.write(row_num, rec_col_idx, rec_val, fmt_expand)
                                elif rec_val == 'Retain':
                                    worksheet.write(row_num, rec_col_idx, rec_val, fmt_retain)
                                elif rec_val == 'Delist':
                                    worksheet.write(row_num, rec_col_idx, rec_val, fmt_delist)
                                else:
                                    worksheet.write(row_num, rec_col_idx, rec_val)

                        for i, col in enumerate(headers):
                            max_len = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
                            worksheet.set_column(i, i, max_len)
                    else:
                        from openpyxl.styles import PatternFill
                        from openpyxl.utils import get_column_letter

                        ws = worksheet
                        headers = df_export.columns.tolist()
                        if 'Recommendation' in headers:
                            rec_col_idx = df_export.columns.get_loc('Recommendation') + 1
                            for row_idx, rec_val in enumerate(df_export['Recommendation'], start=2):
                                cell = ws.cell(row=row_idx, column=rec_col_idx)
                                if rec_val == 'Expand':
                                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                                elif rec_val == 'Retain':
                                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                                elif rec_val == 'Delist':
                                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                        for i, col in enumerate(headers, start=1):
                            max_len = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
                            ws.column_dimensions[get_column_letter(i)].width = max_len

                output.seek(0)
                return output.getvalue()
            except Exception:
                output.close()
                continue
        st.warning("Excel export unavailable: please install `xlsxwriter` or `openpyxl`.")
        return None

    st.subheader("⬇️ Download Results")
    csv = df_export.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="Download as CSV",
        data=csv,
        file_name="SKU_Recommendations.csv",
        mime="text/csv"
    )

    excel_data = to_excel_bytes(df_export)
    if excel_data:
        st.download_button(
            label="Download color-coded Excel (.xlsx)",
            data=excel_data,
            file_name="SKU_Recommendations.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Color-coded Excel export is disabled. Install xlsxwriter with: `pip install xlsxwriter`. ")
